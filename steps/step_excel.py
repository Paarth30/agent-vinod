"""
Excel job tracker — one file, two sheets.
Sheet 1: Discovered Jobs — rebuilt from discovered_jobs.json each run; duplicates collapsed
         into one row with a "Visits" counter.
Sheet 2: Applications & Status — append-only log of actions taken.
"""
import json
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("Run: pip install openpyxl")

from steps.step1_discover import _job_key

TRACKER_PATH = Path("data/job_tracker.xlsx")

# ── Column definitions ─────────────────────────────────────────────────────────
# Sheet 1 — 17 columns
JOBS_COLS = [
    ("Date Found",    16),
    ("Posted",        14),   # "X days ago" from LinkedIn
    ("Visits",         9),   # times the same job appeared across runs
    ("Title",         35),
    ("Company",       22),
    ("Location",      22),
    ("Work Type",     12),
    ("Priority",      10),
    ("ATS Score",     12),
    ("ATS Label",     12),
    ("Skills /35",    11),
    ("Keywords /30",  12),
    ("Exp /20",       10),
    ("Edu /10",       10),
    ("Status",        16),
    ("Link",          40),
    ("JD (excerpt)",  60),
]

# Column indices (1-based) — keep as constants so they never silently drift
COL_DATE      = 1
COL_POSTED    = 2
COL_VISITS    = 3
COL_TITLE     = 4
COL_COMPANY   = 5
COL_LOCATION  = 6
COL_WORKTYPE  = 7
COL_PRIORITY  = 8
COL_ATS_SCORE = 9
COL_ATS_LABEL = 10
COL_SKILLS    = 11
COL_KEYWORDS  = 12
COL_EXP       = 13
COL_EDU       = 14
COL_STATUS    = 15
COL_LINK      = 16
COL_JD        = 17

# Sheet 2 — 9 columns (unchanged)
STATUS_COLS = [
    ("Date",           18),
    ("Company",        22),
    ("Title",          35),
    ("Status",         18),
    ("Method",         20),
    ("Contact Name",   22),
    ("Contact Email",  28),
    ("Resume Used",    28),
    ("Notes",          40),
]

# ── Palette ────────────────────────────────────────────────────────────────────
HEADER_BG = "1A2D3E"
HEADER_FG = "FFFFFF"
ROW_ODD   = "F0F5FA"
ROW_EVEN  = "FFFFFF"

STATUS_COLORS = {
    "New":             ("EEF2FF", "3730A3"),
    "Applied":         ("DCFCE7", "166534"),
    "Saved for Later": ("DBEAFE", "1E40AF"),
    "Rejected":        ("FEE2E2", "991B1B"),
}

TYPE_COLORS = {
    "remote":  ("DCFCE7", "166534"),
    "hybrid":  ("FEF9C3", "854D0E"),
    "on-site": ("FEE2E2", "991B1B"),
}
ATS_COLORS = {
    "Excellent": ("DCFCE7", "166534"),
    "Good":      ("D1FAE5", "065F46"),
    "Fair":      ("FEF9C3", "854D0E"),
    "Low":       ("FEE2E2", "991B1B"),
}


# ── Helpers ────────────────────────────────────────────────────────────────────

def _header_style():
    return {
        "font":      Font(bold=True, color=HEADER_FG, size=10),
        "fill":      PatternFill("solid", fgColor=HEADER_BG),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
    }


def _apply(cell, **kwargs):
    for attr, val in kwargs.items():
        setattr(cell, attr, val)


def _setup_sheet(ws, cols):
    ws.row_dimensions[1].height = 30
    for i, (name, width) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i, value=name)
        _apply(cell, **_header_style())
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"


def _get_or_create_wb():
    if TRACKER_PATH.exists():
        return openpyxl.load_workbook(TRACKER_PATH)
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    _setup_sheet(wb.create_sheet("Discovered Jobs"),       JOBS_COLS)
    _setup_sheet(wb.create_sheet("Applications & Status"), STATUS_COLS)
    return wb


def _row_fill(row_num):
    return PatternFill("solid", fgColor=ROW_ODD if row_num % 2 == 1 else ROW_EVEN)


def _write_job_row(ws, row_num, job):
    ats       = job.get("ats") or {}
    score     = ats.get("score", 0)
    label     = ats.get("label", "N/A")
    breakdown = ats.get("breakdown") or {}
    work_type = (job.get("work_type") or "").lower()
    status    = job.get("status", "New")
    visits    = job.get("visit_count", 1)

    values = [
        (job.get("discovered_at") or datetime.now().isoformat())[:10],  # Date Found
        job.get("posted_text", ""),                                       # Posted
        visits,                                                           # Visits
        job.get("title", ""),
        job.get("company", ""),
        job.get("location", ""),
        work_type.capitalize(),
        job.get("priority", ""),
        f"{score}%" if score else "N/A",
        label,
        breakdown.get("skills", ""),
        breakdown.get("keywords", ""),
        breakdown.get("experience", ""),
        breakdown.get("education", ""),
        status,
        job.get("link", ""),
        (job.get("jd") or "")[:500],
    ]

    bg = _row_fill(row_num)
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.fill = bg
        cell.alignment = Alignment(vertical="top", wrap_text=(col == COL_JD))
        cell.font = Font(size=9)

        if col == COL_WORKTYPE and work_type in TYPE_COLORS:
            pill_bg, pill_fg = TYPE_COLORS[work_type]
            cell.fill = PatternFill("solid", fgColor=pill_bg)
            cell.font = Font(size=9, bold=True, color=pill_fg)

        elif col == COL_ATS_LABEL and label in ATS_COLORS:
            pill_bg, pill_fg = ATS_COLORS[label]
            cell.fill = PatternFill("solid", fgColor=pill_bg)
            cell.font = Font(size=9, bold=True, color=pill_fg)

        elif col == COL_STATUS and status in STATUS_COLORS:
            pill_bg, pill_fg = STATUS_COLORS[status]
            cell.fill = PatternFill("solid", fgColor=pill_bg)
            cell.font = Font(size=9, bold=True, color=pill_fg)

        elif col == COL_VISITS and visits > 1:
            # Highlight frequently-seen jobs
            cell.fill = PatternFill("solid", fgColor="FEF9C3")
            cell.font = Font(size=9, bold=True, color="854D0E")
            cell.alignment = Alignment(horizontal="center", vertical="top")

        elif col == COL_LINK and val:
            cell.hyperlink = str(val)
            cell.font = Font(size=9, color="2563EB", underline="single")

    ws.row_dimensions[row_num].height = 40


def _write_status_row(ws, row_num, job):
    status = job.get("status", "Applied")
    values = [
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        job.get("company", ""),
        job.get("title", ""),
        status,
        job.get("apply_method", ""),
        (job.get("contact") or {}).get("name", ""),
        (job.get("contact") or {}).get("email", ""),
        job.get("resume_docx", ""),
        job.get("notes", ""),
    ]
    bg = _row_fill(row_num)
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.fill = bg
        cell.alignment = Alignment(vertical="top")
        cell.font = Font(size=9)
        if col == 4 and status in STATUS_COLORS:
            pill_bg, pill_fg = STATUS_COLORS[status]
            cell.fill = PatternFill("solid", fgColor=pill_bg)
            cell.font = Font(size=9, bold=True, color=pill_fg)
    ws.row_dimensions[row_num].height = 18


def _is_current_format(ws) -> bool:
    """Return True if the sheet already has the 17-column layout."""
    header = [ws.cell(row=1, column=i).value for i in range(1, 18)]
    return header[COL_VISITS - 1] == "Visits" and header[COL_LINK - 1] == "Link"


# ── Public API ─────────────────────────────────────────────────────────────────

def update_from_discovered(discovered_json: Path = Path("data/discovered_jobs.json")):
    """
    Rebuild 'Discovered Jobs' from discovered_jobs.json.
    Duplicates (same link) are collapsed into one row; the Visits counter
    reflects how many times that job appeared across all runs.
    Any Status edits the user made in Excel are preserved (new-format sheet only).
    """
    if not discovered_json.exists():
        print("No discovered_jobs.json found.")
        return 0

    with open(discovered_json, encoding="utf-8") as f:
        all_records = json.load(f)

    link_order = []
    link_map   = {}
    for record in all_records:
        key = _job_key(record)
        if not key:
            continue
        if key not in link_map:
            link_order.append(key)
            link_map[key] = {"count": 0, "job": record}
        link_map[key]["count"] += 1
        link_map[key]["job"] = record  # keep most recent metadata

    TRACKER_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = _get_or_create_wb()
    ws = wb["Discovered Jobs"]

    # Cache once — used in both preservation and migration branches below
    is_current = _is_current_format(ws)

    # Preserve user's Status edits if the sheet is already in the new format
    saved_status: dict[str, str] = {}
    if is_current:
        for row in ws.iter_rows(min_row=2, values_only=True):
            status = row[COL_STATUS - 1]
            if not status:
                continue
            job = {
                "link":    row[COL_LINK - 1] or "",
                "title":   row[COL_TITLE - 1] or "",
                "company": row[COL_COMPANY - 1] or "",
            }
            saved_status[_job_key(job)] = str(status)

    # Rebuild header if migrating from old format; otherwise just clear data rows
    if not is_current:
        ws.delete_rows(1, ws.max_row)   # removes values AND styling — no ghost rows
        _setup_sheet(ws, JOBS_COLS)
    else:
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

    # Write deduplicated rows
    for row_num, key in enumerate(link_order, 2):
        info = link_map[key]
        job  = {
            **info["job"],
            "visit_count": info["count"],
            "status": saved_status.get(key, info["job"].get("status", "New")),
        }
        _write_job_row(ws, row_num, job)

    try:
        wb.save(TRACKER_PATH)
    except PermissionError:
        print("[!] Cannot save job_tracker.xlsx — close the file in Excel first, then re-run.")
        return 0

    total = len(link_order)
    dupes = len(all_records) - total
    print(f"Excel updated — {total} unique job(s)" + (f", {dupes} duplicate(s) collapsed." if dupes else "."))
    return total


def get_job_statuses() -> dict:
    """Read the current 'Discovered Jobs' sheet and return {job_key: status}.
    Used to exclude already-Applied/Rejected jobs from a fresh search before they're re-processed."""
    if not TRACKER_PATH.exists():
        return {}
    wb = openpyxl.load_workbook(TRACKER_PATH)
    if "Discovered Jobs" not in wb.sheetnames:
        return {}
    ws = wb["Discovered Jobs"]
    if not _is_current_format(ws):
        return {}

    statuses = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        status = row[COL_STATUS - 1]
        if not status:
            continue
        job = {
            "link":    row[COL_LINK - 1] or "",
            "title":   row[COL_TITLE - 1] or "",
            "company": row[COL_COMPANY - 1] or "",
        }
        statuses[_job_key(job)] = str(status)
    return statuses


def get_title_company_statuses() -> dict:
    """Same as get_job_statuses(), but keyed by (title, company) instead of the
    LinkedIn job ID. Some recruiters repost the exact same listing under a brand
    new job ID every day or two — a plain ID-based skip check never recognizes
    those as "already handled," so a job you rejected keeps resurfacing under a
    different ID. Matching on title+company as well catches that."""
    if not TRACKER_PATH.exists():
        return {}
    wb = openpyxl.load_workbook(TRACKER_PATH)
    if "Discovered Jobs" not in wb.sheetnames:
        return {}
    ws = wb["Discovered Jobs"]
    if not _is_current_format(ws):
        return {}

    statuses = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        status = row[COL_STATUS - 1]
        if not status:
            continue
        title = str(row[COL_TITLE - 1] or "").strip().lower()
        company = str(row[COL_COMPANY - 1] or "").strip().lower()
        if not (title and company):
            continue
        key = (title, company)
        # A repost gets its own untouched "New" row alongside an older
        # Applied/Rejected one for the same title+company — once ever acted
        # on, that verdict should win regardless of which row comes last.
        if str(status) in ("Applied", "Rejected") or key not in statuses:
            statuses[key] = str(status)
    return statuses


def log_application(job: dict):
    """Append a row to Applications & Status when a job is applied/actioned."""
    TRACKER_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = _get_or_create_wb()
    ws = wb["Applications & Status"]
    _write_status_row(ws, ws.max_row + 1, job)
    try:
        wb.save(TRACKER_PATH)
    except PermissionError:
        print("[!] Cannot save job_tracker.xlsx — close the file in Excel first.")


def update_job_status(link: str, new_status: str):
    """Update the Status cell in Discovered Jobs for a given link."""
    if not TRACKER_PATH.exists():
        return
    wb = openpyxl.load_workbook(TRACKER_PATH)
    ws = wb["Discovered Jobs"]
    for row in ws.iter_rows(min_row=2):
        if str(row[COL_LINK - 1].value or "").strip() == link.strip():
            cell = row[COL_STATUS - 1]
            cell.value = new_status
            if new_status in STATUS_COLORS:
                pill_bg, pill_fg = STATUS_COLORS[new_status]
                cell.fill = PatternFill("solid", fgColor=pill_bg)
                cell.font = Font(size=9, bold=True, color=pill_fg)
            break
    wb.save(TRACKER_PATH)


if __name__ == "__main__":
    update_from_discovered()
    print(f"Tracker saved to: {TRACKER_PATH.resolve()}")
